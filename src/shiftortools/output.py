"""Output helpers: write JSON and Excel results."""
import json
from pathlib import Path
from typing import Any, Dict
import openpyxl
from openpyxl.utils import get_column_letter


def write_json(obj: Any, out_path: str):
    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open('w', encoding='utf-8') as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def write_excel(shiftjson: Dict[str, Any], solver_result: Dict[str, Any], out_path: str):
    """Create a simple Excel workbook with sheets: shift_result, ng_calendar, unknown_names

    - `shift_result`: rows date | hospital | assigned names (comma-separated)
    - `ng_calendar`: person | ng_dates (comma-separated)
    - `unknown_names`: list
    """
    wb = openpyxl.Workbook()

    # schedule sheet: create layout matching existing template
    from datetime import datetime
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.active
    ws.title = 'schedule'

    # Title row (A1 merged)
    month_title = solver_result.get('month') or shiftjson.get('month') or ''
    if not month_title:
        # try derive from dates
        dates = solver_result.get('dates', [])
        if dates:
            try:
                dt = datetime.fromisoformat(dates[0])
                month_title = f"{dt.year}年{dt.month}月分"
            except Exception:
                month_title = ''
    # Do not merge title across columns; keep in A1 only
    ws['A1'] = month_title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Subtitle / description
    # Do not merge subtitle; keep in A2 only
    ws['A2'] = '院外研修、大学救急研修・県内協力病院二次救急輪番研修（担当表）'
    ws['A2'].alignment = Alignment(horizontal='center')

    # Header rows (row 4 area)
    header_row = 4
    # style header
    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='FFDCE6F1')
    thin = Side(border_style='thin', color='FFBBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 12):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # prepare specific header/subheader/data fills per hospital groups
    # choose colors approximating the provided screenshot
    def hex_fill(hexcode):
        h = hexcode.lstrip('#')
        if len(h) == 6:
            return PatternFill('solid', fgColor='FF' + h.upper())
        return PatternFill('solid', fgColor=h.upper())

    # Group 1 (D,E) - university
    g1_hdr = hex_fill('E8B8B6')
    g1_sub = hex_fill('F4DFDE')
    # Group 2 (F,G) - iwasaki
    g2_hdr = hex_fill('D7EACD')
    g2_sub = hex_fill('EEF7E9')
    # Group 3 (H,I) - nagai
    g3_hdr = hex_fill('D9CFE6')
    g3_sub = hex_fill('EFEAF6')
    # Group 4 (J,K) - toyama
    g4_hdr = hex_fill('F5D6C2')
    g4_sub = hex_fill('FCEDE4')

    # override header fills for each hospital group (row4)
    for c in (4, 5):
        ws.cell(row=header_row, column=c).fill = g1_hdr
    for c in (6, 7):
        ws.cell(row=header_row, column=c).fill = g2_hdr
    for c in (8, 9):
        ws.cell(row=header_row, column=c).fill = g3_hdr
    for c in (10, 11):
        ws.cell(row=header_row, column=c).fill = g4_hdr

    # Place requested labels and merges:
    # Merge A,B,C across rows 4-6 (leftmost info block)
    ws.merge_cells(start_row=4, start_column=1, end_row=6, end_column=1)
    ws.merge_cells(start_row=4, start_column=2, end_row=6, end_column=2)
    ws.merge_cells(start_row=4, start_column=3, end_row=6, end_column=3)
    ws.cell(row=4, column=3).value = '院外研修（献血・KKC・MHMC）'
    ws.cell(row=4, column=3).alignment = Alignment(horizontal='center', vertical='center')

    # Remove background fill for the merged A-C block (rows 4-6)
    for rr in range(4, 7):
        for cc in (1, 2, 3):
            ws.cell(row=rr, column=cc).fill = PatternFill(fill_type=None)

    # L4:L6 merge for 備考
    ws.merge_cells(start_row=4, start_column=12, end_row=6, end_column=12)
    ws.cell(row=4, column=12).value = '備考'
    ws.cell(row=4, column=12).alignment = Alignment(horizontal='center', vertical='center')

    # Merge hospital label pairs: D4:E4, F4:G4, H4:I4, J4:K4 and set labels (main headers on row4)
    ws.merge_cells(start_row=header_row, start_column=4, end_row=header_row, end_column=5)
    ws.cell(row=header_row, column=4).value = '大学病院'
    ws.merge_cells(start_row=header_row, start_column=6, end_row=header_row, end_column=7)
    ws.cell(row=header_row, column=6).value = '岩崎病院'
    ws.merge_cells(start_row=header_row, start_column=8, end_row=header_row, end_column=9)
    ws.cell(row=header_row, column=8).value = '永井病院'
    ws.merge_cells(start_row=header_row, start_column=10, end_row=header_row, end_column=11)
    ws.cell(row=header_row, column=10).value = '遠山病院'

    # ensure center alignment and header font for the main headers
    for col_idx in (4,6,8,10):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Insert subheader rows (row5 and row6) with merges and texts per user request
    # Merge D5:E6 into one block and put the subheader text
    ws.merge_cells(start_row=5, start_column=4, end_row=6, end_column=5)
    d5 = ws.cell(row=5, column=4)
    d5.value = '準夜：16:30～23:30\n日勤(土・日・祝)：8:30～17:15'
    d5.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')
    # apply subheader fill for merged D5:E6
    for rr in (5, 6):
        for cc in (4, 5):
            ws.cell(row=rr, column=cc).fill = g1_sub

    # F5:G5 and F6:G6
    ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=7)
    f5 = ws.cell(row=5, column=6)
    f5.value = '(火)18:30～23:30'
    f5.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')
    # F5:F6 fills
    for cc in (6, 7):
        ws.cell(row=5, column=cc).fill = g2_sub
        ws.cell(row=6, column=cc).fill = g2_sub
    ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7)
    f6 = ws.cell(row=6, column=6)
    f6.value = '(内科)'
    f6.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')

    # H5:I5 and H6:I6
    ws.merge_cells(start_row=5, start_column=8, end_row=5, end_column=9)
    h5 = ws.cell(row=5, column=8)
    h5.value = '(土)日勤：8:30～16:30\n(土)準夜：16:00～23:30'
    h5.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')
    # H5:H6 fills
    for cc in (8, 9):
        ws.cell(row=5, column=cc).fill = g3_sub
        ws.cell(row=6, column=cc).fill = g3_sub
    ws.merge_cells(start_row=6, start_column=8, end_row=6, end_column=9)
    h6 = ws.cell(row=6, column=8)
    h6.value = '(内科：日勤、準夜)'
    h6.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')

    # J5:K6 merged with multiline schedule (merge J&K, rows 5-6)
    ws.merge_cells(start_row=5, start_column=10, end_row=6, end_column=11)
    j5 = ws.cell(row=5, column=10)
    j5.value = '(火)18:00～23:30\n(土)14:00～22:00\n(日)（祝）17:00～23:30'
    j5.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')
    # J5:K6 fills (subheader/data background)
    for rr in (5, 6):
        for cc in (10, 11):
            ws.cell(row=rr, column=cc).fill = g4_sub

    # Set row heights for rows 5 and 6 as requested
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 30

    # map hospitals to columns per user's description
    hosp_cols = {'大学病院': 5, '岩崎病院': 7, '永井病院': 9, '遠山病院': 10}
    # offsite entries saved in shiftjson (date_iso -> [raw strings])
    offsite_map = shiftjson.get('offsite_entries', {}) if shiftjson else {}

    # Load hospital weekday/date slots config if available to determine pre-configured slots
    from pathlib import Path
    cfg = {}
    cfg_path = Path('config/hospital_weekday_slots.json')
    if cfg_path.exists():
        try:
            with cfg_path.open('r', encoding='utf-8') as cf:
                raw_cfg = json.load(cf)
            # normalize to {hospital: {key_str: int}}
            for h, m in raw_cfg.items():
                nm = {}
                if isinstance(m, dict):
                    for k, v in m.items():
                        try:
                            nm[str(k)] = int(v)
                        except Exception:
                            try:
                                nm[str(k)] = int(float(v))
                            except Exception:
                                nm[str(k)] = 0
                cfg[h] = nm
        except Exception:
            cfg = {}

    def hospital_has_slot_for_date(hospital_name: str, date_obj) -> bool:
        """Return True if hospital has configured slots for date (either exact date key or weekday)."""
        if not cfg or hospital_name not in cfg:
            return False
        m = cfg.get(hospital_name, {})
        ds = date_obj.isoformat()
        if ds in m:
            try:
                return int(m[ds]) > 0
            except Exception:
                return False
        wk = str(date_obj.weekday())
        try:
            return int(m.get(wk, 0)) > 0
        except Exception:
            return False

    # iterate dates and write blocks
    # data starts after header rows 4-6
    start_row = header_row + 3
    dates = solver_result.get('dates', [])
    assignments = solver_result.get('assignments', {})
    from datetime import date as _date
    from src.shiftortools.utils import is_holiday

    cur_row = start_row
    # map each output row to its date string (for border decisions)
    row_date_map = {}
    for dstr in dates:
        try:
            dd = datetime.fromisoformat(dstr).date()
        except Exception:
            # skip unparsable
            continue
        wd = dd.weekday()  # Mon=0
        # determine rows per date: weekend/holiday -> 4, else max(2, offsite_count)
        is_weekend = wd >= 5 or is_holiday(dd)
        # filter out empty offsite entries
        offsite_entries = [s for s in offsite_map.get(dstr, []) if str(s).strip()]
        # determine max assignments per hospital for this date
        max_assigned = 0
        for h in hosp_cols.keys():
            assigned = assignments.get(dstr, {}).get(h, [])
            if assigned:
                max_assigned = max(max_assigned, len(assigned))

        if is_weekend:
            rows_per_date = 4
        else:
            # ensure rows accommodate at least the largest per-hospital stack and offsite entries
            rows_per_date = max(2, max_assigned, len(offsite_entries))

        # write date and weekday in first column(s) spanning the block
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row+rows_per_date-1, end_column=1)
        ws.cell(row=cur_row, column=1).value = f"{dd.day}日"
        ws.cell(row=cur_row, column=1).alignment = Alignment(vertical='top', horizontal='left')

        ws.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row+rows_per_date-1, end_column=2)
        weekday_label = ['月','火','水','木','金','土','日'][wd]
        bcell = ws.cell(row=cur_row, column=2)
        bcell.value = weekday_label
        bcell.alignment = Alignment(vertical='top', horizontal='center')
        # if holiday (national) or weekend, color weekday red
        if is_holiday(dd) or wd >= 5:
            bcell.font = Font(color='FF0000')

        # Fill offsite C column: join entries with newlines across the rows (one per line)
        if offsite_entries:
            # put entries in consecutive rows of column C
            for i, val in enumerate(offsite_entries):
                if i >= rows_per_date:
                    # append extras into last row
                    existing = ws.cell(row=cur_row+rows_per_date-1, column=3).value or ''
                    newline = (existing + '\n' if existing else '') + val
                    ws.cell(row=cur_row+rows_per_date-1, column=3).value = newline
                else:
                    ws.cell(row=cur_row+i, column=3).value = val

        # For each hospital, place assigned names into rows (one per cell top-down)
        for h, col_idx in hosp_cols.items():
            assigned = assignments.get(dstr, {}).get(h, [])
            for i in range(rows_per_date):
                if i < len(assigned):
                    ws.cell(row=cur_row+i, column=col_idx).value = assigned[i]
                else:
                    ws.cell(row=cur_row+i, column=col_idx).value = None

        # Determine per-hospital white background rules:
        # - 大学病院: white for all dates
        # - other hospitals: white only on dates where slots are assigned
        uni_white = True
        iwa_white = bool(assignments.get(dstr, {}).get('岩崎病院'))
        nagai_white = bool(assignments.get(dstr, {}).get('永井病院'))
        toyo_white = bool(assignments.get(dstr, {}).get('遠山病院'))

        # Weekend special: for university hospital (D column) merge top2 and bottom2 rows and label 日勤/準夜
        if rows_per_date == 4 and (wd >= 5 or is_holiday(dd)):
            # merge top two rows for D
            ws.merge_cells(start_row=cur_row, start_column=4, end_row=cur_row+1, end_column=4)
            ws.cell(row=cur_row, column=4).value = '日勤'
            ws.cell(row=cur_row, column=4).alignment = Alignment(horizontal='center', vertical='top')
            # merge bottom two rows for D
            ws.merge_cells(start_row=cur_row+2, start_column=4, end_row=cur_row+3, end_column=4)
            ws.cell(row=cur_row+2, column=4).value = '準夜'
            ws.cell(row=cur_row+2, column=4).alignment = Alignment(horizontal='center', vertical='top')

            # Also apply same merge/labels for 永井病院 (H column)
            # Only merge/label H (永井病院) when that hospital has pre-configured slots for this date
            try:
                if hospital_has_slot_for_date('永井病院', dd):
                    ws.merge_cells(start_row=cur_row, start_column=8, end_row=cur_row+1, end_column=8)
                    ws.cell(row=cur_row, column=8).value = '日勤'
                    ws.cell(row=cur_row, column=8).alignment = Alignment(horizontal='center', vertical='top')
                    ws.merge_cells(start_row=cur_row+2, start_column=8, end_row=cur_row+3, end_column=8)
                    ws.cell(row=cur_row+2, column=8).value = '準夜'
                    ws.cell(row=cur_row+2, column=8).alignment = Alignment(horizontal='center', vertical='top')
            except Exception:
                # if config check fails, skip the H-column merge to be safe
                pass

        # For iwasaki (岩崎病院) in 2-row blocks, if there are assignments, merge F column two rows and label 準夜
        if rows_per_date == 2:
            iw_assigned = assignments.get(dstr, {}).get('岩崎病院', [])
            if iw_assigned:
                ws.merge_cells(start_row=cur_row, start_column=6, end_row=cur_row+1, end_column=6)
                ws.cell(row=cur_row, column=6).value = '準夜'
                ws.cell(row=cur_row, column=6).alignment = Alignment(horizontal='center', vertical='top')

        # apply column background fills for the block rows using group sub fills
        for rr in range(cur_row, cur_row+rows_per_date):
            # record row->date mapping for border decisions
            row_date_map[rr] = dstr
            # Decide per-group whether to use white or subcolor for this date
            white = PatternFill('solid', fgColor='FFFFFFFF')
            # Group1 (大学): always white
            if uni_white:
                ws.cell(row=rr, column=4).fill = white
                ws.cell(row=rr, column=5).fill = white
            else:
                ws.cell(row=rr, column=4).fill = g1_sub
                ws.cell(row=rr, column=5).fill = g1_sub

            # Group2 (岩崎): white if iwa_white else sub
            if iwa_white:
                ws.cell(row=rr, column=6).fill = white
                ws.cell(row=rr, column=7).fill = white
            else:
                ws.cell(row=rr, column=6).fill = g2_sub
                ws.cell(row=rr, column=7).fill = g2_sub

            # Group3 (永井): white if nagai_white else sub
            if nagai_white:
                ws.cell(row=rr, column=8).fill = white
                ws.cell(row=rr, column=9).fill = white
            else:
                ws.cell(row=rr, column=8).fill = g3_sub
                ws.cell(row=rr, column=9).fill = g3_sub

            # Group4 (遠山): white if toyo_white else sub
            if toyo_white:
                ws.cell(row=rr, column=10).fill = white
                ws.cell(row=rr, column=11).fill = white
            else:
                ws.cell(row=rr, column=10).fill = g4_sub
                ws.cell(row=rr, column=11).fill = g4_sub

            # For 遠山 (J/K), if that specific row has assignment(s), merge J and K horizontally for that row
            # Note: assignments were already written into columns 10/11 above
            jval = (ws.cell(row=rr, column=10).value or '')
            kval = (ws.cell(row=rr, column=11).value or '')
            if str(jval).strip() or str(kval).strip():
                try:
                    ws.merge_cells(start_row=rr, start_column=10, end_row=rr, end_column=11)
                except Exception:
                    pass

            # Merge L column vertically for this date block (per-day notes column)
            try:
                ws.merge_cells(start_row=cur_row, start_column=12, end_row=cur_row+rows_per_date-1, end_column=12)
            except Exception:
                pass

        cur_row += rows_per_date

    last_row = cur_row - 1

    # Draw grid borders from the header row (row4) to last_row for columns A-L
    # Principles:
    # - Default: thin solid grid for all cells A-L between row4..last_row
    # - Within the same date block (multiple rows for one date): horizontal separators
    #   are dotted ONLY for columns C,E,G,I,K (3,5,7,9,11). Other columns keep solid lines.
    # - Exception (no border): for non-university hospital slot pairs (F/G, H/I, J/K => (6,7),(8,9),(10,11))
    #   when both cells in the pair are empty for that row, remove all borders for those two cells.
    thin_side = Side(border_style='thin', color='FF000000')
    dotted_side = Side(border_style='dotted', color='FF000000')
    no_side = Side(border_style=None)
    dotted_cols = {3, 5, 7, 9, 11}
    non_uni_col_range = set(range(6, 12))  # F(6) .. K(11)
    # hospital pairs left/right mapping for internal vertical removal
    pair_map = {6: (6,7), 7: (6,7), 8: (8,9), 9: (8,9), 10: (10,11), 11: (10,11)}
    non_uni_pairs = [(6, 7), (8, 9), (10, 11)]

    if last_row >= header_row:
        for r in range(header_row, last_row + 1):
            prev_date = row_date_map.get(r-1)
            this_date = row_date_map.get(r)
            next_date = row_date_map.get(r+1)
            same_above = prev_date is not None and this_date is not None and prev_date == this_date
            same_below = this_date is not None and next_date is not None and this_date == next_date

            for c in range(1, 13):
                # default vertical borders
                left_side = thin_side
                right_side = thin_side

                # For non-university columns F-K:
                # - remove internal horizontal borders within the same date (top/bottom = none)
                # - remove internal vertical border between the pair columns (e.g., F/G),
                #   but keep vertical separators at hospital boundaries (between pairs)
                if c in non_uni_col_range:
                    # horizontal: remove intra-date (both top and bottom none), keep date-boundary thin
                    top_side = no_side if same_above else thin_side
                    bottom_side = no_side if same_below else thin_side

                    # vertical: remove internal vertical between pair columns (e.g., F/G), keep outer separators thin
                    a, b = pair_map[c]
                    if c == a:
                        # left boundary of pair stays thin, right internal removed
                        left_side = thin_side
                        right_side = no_side
                    else:
                        # right boundary of pair stays thin, left internal removed
                        left_side = no_side
                        right_side = thin_side
                else:
                    # horizontal borders: dotted only for dotted_cols when within same date
                    top_side = dotted_side if (same_above and c in dotted_cols) else thin_side
                    bottom_side = dotted_side if (same_below and c in dotted_cols) else thin_side

                border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)
                ws.cell(row=r, column=c).border = border

    # Styling: column widths and freeze pane
    # Set column widths per user request:
    # A:7.17, B:5.33, C:30.33,
    # D,F,H,J:5.17 and E,G,I,K:30.33
    col_widths = {
        1: 7.17,
        2: 5.33,
        3: 30.33,
        4: 5.17,
        5: 30.33,
        6: 5.17,
        7: 30.33,
        8: 5.17,
        9: 30.33,
        10: 5.17,
        11: 30.33,
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = ws['A7']

    # ng_calendar
    w2 = wb.create_sheet('ng_calendar')
    w2.append(['name', 'ng_dates', 'ng_reasons'])
    for r in shiftjson.get('residents', []):
        name = r.get('name')
        ng_dates = r.get('ng_dates', [])
        ng_reasons = r.get('ng_reasons', {})
        # flatten reasons to short string
        reasons = []
        for d in ng_dates:
            reasons.append(f"{d}:{'|'.join(ng_reasons.get(d, []))}")
        w2.append([name, ', '.join(ng_dates), '; '.join(reasons)])

    # unknown_names
    w3 = wb.create_sheet('unknown_names')
    w3.append(['unknown_name'])
    for n in shiftjson.get('unknown_names', []):
        w3.append([n])

    # violations placeholder
    w4 = wb.create_sheet('violations')
    w4.append(['note'])

    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(p))
